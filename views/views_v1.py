import csv
import io
import json
import logging
import re
import shlex

import openpyxl
from flask import Blueprint, request, Response, send_file, session, jsonify
from openpyxl.utils import get_column_letter

from src import constants
from src.constants import SLASH_COMMAND
from src.models.form import SlackForm, Submission
from src.services import submissions_service, forms_service, schedule_management_service
from src.services.pagination_service import handle_pagination
from src.slack_api.slack_user import SlackUser
from src.slack_ui import responses as slack_ui_responses
from views.auth import user_logged_in_api
from views.view_utils import verify_slack_request, form_visible_to_user

urls_v1 = Blueprint('views_v1', __name__)


@urls_v1.route("/slash-command", methods=['POST'])
@verify_slack_request
def slack_webhook():
    user = SlackUser(user_id=request.form['user_id'])
    response_url = request.form['response_url']
    command_text = (request.form['text'] or '').replace("”", '"').replace("“", '"')
    command_text = re.sub(r'\s*=\s*', '=', command_text)  # Remove spaces before and after the equal sign
    args = shlex.split(command_text)
    logging.info(f"[{user.username}] from [{user.team_id}] called /{SLASH_COMMAND} {command_text}")
    result = None
    if len(args) < 1:
        logging.info(f"[{user.username}] from [{user.team_id}] - Returning help text block")
        result = slack_ui_responses.help_text_response
    elif args[0] == "create":
        logging.info(f"[{user.username}] from [{user.team_id}] - Trying to create form")
        result = forms_service.create_form_command(user, args[1:], command_text, response_url)
    elif args[0] == "list":
        result = forms_service.list_forms_command(user, response_url)
    elif args[0] == "fill":
        logging.info(f"[{user.username}] from [{user.team_id}] - Trying to fill form")
        result = forms_service.fill_form_command(user, args[1:], response_url)
    if result:
        return Response(response=json.dumps(result), status=200, mimetype="application/json")
    return Response(status=200, mimetype="application/json")


@urls_v1.route("/interactive", methods=['POST'])
@verify_slack_request
def slack_interactive_endpoint():
    payload = json.loads(request.form['payload'])
    user = SlackUser(user_id=payload['user']['id'])
    response_url = payload['response_url']
    result = None
    for action in payload['actions']:
        action_id = action['action_id']
        if action_id in (constants.FORM_WEEKDAYS, constants.FORM_TIME, constants.VIEW_FORM_SUBMISSIONS) \
                or action.get('type') == 'static_select':
            return Response(status=200, mimetype="application/json")
        value = action['value']
        if action_id == constants.DELETE_FORM:
            result = forms_service.delete_form_command(value, user, response_url)
        elif action_id == constants.FILL_FORM_NOW:
            result = forms_service.fill_form_now_command(value, response_url)
        elif action_id == constants.SCHEDULE_FORM:
            result = schedule_management_service.schedule_form_command(value, user, response_url)
        elif action_id == constants.CREATE_FORM_SCHEDULE:
            schedule_form_state = payload['state']['values']
            result = schedule_management_service.create_form_schedule_command(value, user, schedule_form_state, response_url)
        elif action_id == constants.DELETE_SCHEDULE:
            result = schedule_management_service.delete_schedule_command(value, user, response_url)
        elif action_id == constants.SUBMIT_FORM_SCHEDULED:
            result = submissions_service.submit_scheduled_form(value, user, payload, response_url)
        elif action_id == constants.SUBMIT_FORM_NOW:
            result = submissions_service.submit_form_now(value, user, payload, response_url)

        elif action_id.startswith(constants.LIST_FORMS_PREFIX) or action_id.startswith(constants.FILL_FORMS_PREFIX):
            prefix = constants.LIST_FORMS_PREFIX if action_id.startswith(constants.LIST_FORMS_PREFIX) else constants.FILL_FORMS_PREFIX
            page_button = action_id[len(prefix):]
            current_page = int(action['block_id'].split(':')[-1])
            result = handle_pagination(prefix, page_button, user, response_url, current_page)

    if result:
        return Response(response=json.dumps(result), status=200, mimetype="application/json")
    return Response(status=200, mimetype="application/json")


@urls_v1.route('/forms', methods=['GET'])
@user_logged_in_api
def forms_view():
    page = int(request.args.get('page', 1))
    per_page = min(int(request.args.get('per_page', 10)), 100)
    team_id = session['user_data']['team_id']
    total = SlackForm.objects(team_id=team_id).count()
    page_forms = SlackForm.objects(team_id=team_id).skip((page - 1) * per_page).limit(per_page)
    forms = [form.to_dict() for form in page_forms]
    return jsonify({
        'forms': forms,
        'page': page,
        'per_page': per_page,
        'total': total,
    })


@urls_v1.route('/submissions', methods=['GET'])
@user_logged_in_api
@form_visible_to_user
def submissions_view():
    # noinspection PyTypeHints
    request.slack_form: SlackForm
    form_id = str(request.slack_form.id)
    form = request.slack_form
    page = int(request.args.get('page', 1))
    per_page = min(int(request.args.get('per_page', 10)), 100)

    total = Submission.objects.filter(form_id=form_id).count()
    page_submissions = Submission.objects.filter(form_id=form_id).skip((page - 1) * per_page).limit(per_page)
    submissions = [submission.to_dict() for submission in page_submissions]

    return jsonify({
        'submissions': submissions,
        'form_id': form_id,
        'form_name': form.name,
        'page': page,
        'per_page': per_page,
        'total': total,
        'SLASH_COMMAND': SLASH_COMMAND,
    })



@urls_v1.route('/submissions/export/csv')
@user_logged_in_api
@form_visible_to_user
def export_submissions_csv():
    # noinspection PyUnresolvedReferences
    form_id = str(request.slack_form.id)
    form = SlackForm.objects.get(id=form_id)
    # noinspection PyShadowingNames
    submissions = Submission.objects(form_id=form_id)

    csv_data = io.StringIO()
    writer = csv.writer(csv_data)
    writer.writerow(['submitted by', 'date'] + [field.title for field in form.fields])  # Columns

    for submission in submissions:
        row = [submission.user_name, f"{submission.formatted_date} {submission.formatted_time}"]
        field_values = {field.title: field.value for field in submission.fields}
        row.extend([field_values.get(field.title, '') for field in form.fields])
        writer.writerow(row)

    csv_data.seek(0)
    csv_bytes = io.BytesIO(csv_data.getvalue().encode())

    return send_file(csv_bytes, mimetype='text/csv', as_attachment=True, download_name=f'{form.name}.csv')


@urls_v1.route('/submissions/export/xlsx')
@user_logged_in_api
@form_visible_to_user
def export_submissions_xlsx():
    # noinspection PyUnresolvedReferences
    form_id = str(request.slack_form.id)
    form = SlackForm.objects.get(id=form_id)
    # noinspection PyShadowingNames
    submissions = Submission.objects(form_id=form_id)

    wb = openpyxl.Workbook()
    ws = wb.active

    # Write headers
    headers = ['submitted by', 'date'] + [field.title for field in form.fields]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header

    # Write data
    for row_num, submission in enumerate(submissions, 2):
        row_data = [submission.user_name, f"{submission.formatted_date} {submission.formatted_time}"]
        field_values = {field.title: field.value for field in submission.fields}
        row_data.extend([field_values.get(field.title, '') for field in form.fields])

        for col_num, cell_value in enumerate(row_data, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}{row_num}'] = cell_value

    xlsx_data = io.BytesIO()
    wb.save(xlsx_data)
    xlsx_data.seek(0)
    return send_file(xlsx_data, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'{form.name}.xlsx')
