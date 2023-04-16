import csv
import io
import json
import logging
import os
import re
import shlex
from datetime import datetime, timedelta

import bson
import math
import openpyxl
import requests
from flask import Flask, request, Response, render_template, session, redirect, url_for, make_response, send_file
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_session import Session
from openpyxl.utils import get_column_letter
from pymongo import MongoClient
from slack_sdk.signature import SignatureVerifier

from src import constants
from src.services import submissions_service, forms_service, schedule_management_service
from src.constants import SLASH_COMMAND
from src.models.connect import connect_to_mongo
from src.models.form import Submission, SlackForm
from src.server_settings import MONGO_DB_URL, CustomRequest, SESSION_COOKIE_NAME, MONGO_DB_NAME, SLACK_CLIENT_ID, \
    SLACK_CLIENT_SECRET, SLACK_OAUTH_URL, SLACK_USER_INFO_URL, DOMAIN
from src.slack_api.slack_user import SlackUser
from src.slack_ui import responses as slack_ui_responses

logging.basicConfig(format='%(asctime)s - %(levelname)s - %(message)s', level=logging.INFO)

connect_to_mongo()
pymongo_client = MongoClient(MONGO_DB_URL)

app = Flask(__name__)
app.request_class = CustomRequest

app.config['SECRET_KEY'] = os.environ['SESSION_KEY']  # Replace with your secret key
app.config['SESSION_TYPE'] = 'mongodb'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=1)
app.config['SESSION_COOKIE_SECURE'] = True
app.config['SESSION_COOKIE_NAME'] = SESSION_COOKIE_NAME
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_MONGODB'] = pymongo_client
app.config['SESSION_MONGODB_DB'] = MONGO_DB_NAME

Session(app)
slack_verifier = SignatureVerifier(os.environ['SIGNING_SECRET'])

limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=["60/minute"],
    storage_uri="memory://",
)


def verify_slack_request(func, *args, **kwargs):
    def wrapper():
        if slack_verifier.is_valid_request(request.get_data(), request.headers):
            return func(*args, **kwargs)
        return Response(status=401)

    wrapper.__name__ = func.__name__
    return wrapper


def user_is_logged_in():
    return 'access_token' in session and 'user_data' in session


def user_logged_in(func, *args, **kwargs):
    def wrapper():
        if request.args.get('code') and 'access_token' not in session or 'user_data' not in session:
            # OAuth2: Exchange the authorization code for an access token
            code = request.args.get('code')
            payload = {
                'client_id': SLACK_CLIENT_ID,
                'client_secret': SLACK_CLIENT_SECRET,
                'code': code
            }
            response = requests.post(SLACK_OAUTH_URL, data=payload)
            response_data = response.json()

            if 'access_token' not in response_data:
                return redirect(url_for('index', redirect_url=request.url))

            # Fetch user information
            auth_token = response_data['access_token']
            user_id = response_data['authed_user']['id']
            headers = {'Authorization': f"Bearer {auth_token}"}
            user_response = requests.get(SLACK_USER_INFO_URL, headers=headers, params={'user': user_id})
            user_data = user_response.json()

            if user_data['ok']:
                session['access_token'] = auth_token
                session['user_data'] = user_data['user']
            else:
                return redirect(url_for('index', redirect_url=request.url))

        if 'access_token' not in session or 'user_data' not in session:
            return redirect(url_for('index', redirect_url=request.url))
        request.user = SlackUser(
            user_id=session['user_data']['id'],
            user_name=session['user_data']['name'],
            team_id=session['user_data']['team_id'],
        )
        return func(*args, **kwargs)

    wrapper.__name__ = func.__name__
    return wrapper


def form_visible_to_user(func, *args, **kwargs):
    def wrapper():
        form_id = request.args.get('formId')
        if not form_id:
            return make_response("You must provide formId query parameter", 400)
        try:
            form = SlackForm.objects.filter(id=form_id).first()
        except bson.errors.InvalidId:
            return make_response("Form not found", 404)
        # noinspection PyUnresolvedReferences
        if form.team_id != request.user.team_id:
            # noinspection PyUnresolvedReferences
            logging.warning(f"User {request.user.id} tried to access a form that doesn't belong to their team")
            return make_response("Form not found", 404)
        request.slack_form = form
        return func(*args, **kwargs)

    wrapper.__name__ = func.__name__
    return wrapper


@app.route('/')
def index():
    if user_is_logged_in():
        return redirect('/forms')
    redirect_path = request.args.get('redirect_path', 'forms')
    return render_template('sign-in.html', SLACK_CLIENT_ID=SLACK_CLIENT_ID, redirect_url=f'{DOMAIN}/{redirect_path}')


@app.route("/logout")
@user_logged_in
def logout():
    if "access_token" in session:
        session.pop("access_token")
    if "user_data" in session:
        session.pop("user_data")
    cookie_name = app.config.get("REMEMBER_COOKIE_NAME", SESSION_COOKIE_NAME)
    if cookie_name in request.cookies:
        session["_remember"] = "clear"
        if "_remember_seconds" in session:
            session.pop("_remember_seconds")
    return redirect('/')


@app.route("/slash-command", methods=['POST'])
@verify_slack_request
def slack_webhook():
    user = SlackUser(
        user_id=request.form['user_id'],
        user_name=request.form['user_name'],
        team_id=request.form['team_id'],
    )
    response_url = request.form['response_url']
    command_text = (request.form['text'] or '').replace("”", '"').replace("“", '"')
    command_text = re.sub(r'\s*=\s*', '=', command_text)  # Remove spaces before and after the equal sign
    args = shlex.split(command_text)
    logging.info(f"[{user.username}] from [{user.team_id}] called /{SLASH_COMMAND} {command_text}")
    result = None
    if len(args) < 1:
        logging.info(f"[{user.username}] from [{user.team_id}] - Returning help text block")
        result = slack_ui_responses.help_text_response
    elif args[0] == "create":
        logging.info(f"[{user.username}] from [{user.team_id}] - Trying to create form")
        result = forms_service.create_form_command(user, args[1:], command_text, response_url)
    elif args[0] == "list":
        result = forms_service.list_forms_command(user, response_url)
    elif args[0] == "fill":
        logging.info(f"[{user.username}] from [{user.team_id}] - Trying to fill form")
        result = forms_service.fill_form_command(user, args[1:], response_url)
    if result:
        return Response(response=json.dumps(result), status=200, mimetype="application/json")
    return Response(status=200, mimetype="application/json")


@app.route("/interactive", methods=['POST'])
@verify_slack_request
def slack_interactive_endpoint():
    payload = json.loads(request.form['payload'])
    user = SlackUser(
        user_id=payload['user']['id'],
        user_name=payload['user']['username'],
        team_id=payload['user']['team_id'],
    )
    response_url = payload['response_url']
    result = None
    for action in payload['actions']:
        action_id = action['action_id']
        if action_id in (constants.FORM_WEEKDAYS, constants.FORM_TIME, constants.VIEW_FORM_SUBMISSIONS) \
                or action.get('type') == 'static_select':
            return Response(status=200, mimetype="application/json")
        value = action['value']
        if action_id == constants.DELETE_FORM:
            result = forms_service.delete_form_command(value, user, response_url)
        elif action_id == constants.FILL_FORM_NOW:
            result = forms_service.fill_form_now_command(value, response_url)
        elif action_id == constants.SCHEDULE_FORM:
            result = schedule_management_service.schedule_form_command(value, response_url)
        elif action_id == constants.CREATE_FORM_SCHEDULE:
            schedule_form_state = payload['state']['values']
            result = schedule_management_service.create_form_schedule_command(value, user, schedule_form_state, response_url)
        elif action_id == constants.DELETE_SCHEDULE:
            result = schedule_management_service.delete_schedule_command(value, user, response_url)
        elif action_id == constants.SUBMIT_FORM_SCHEDULED:
            result = submissions_service.submit_scheduled_form(value, user, payload, response_url)
        elif action_id == constants.SUBMIT_FORM_NOW:
            result = submissions_service.submit_form_now(value, user, payload, response_url)
        elif action_id == constants.LIST_FORMS_PREVIOUS_PAGE:
            current_page = int(action['block_id'].split(':')[-1])  # Extract the current page from the block_id
            result = forms_service.list_forms_command(user, response_url, current_page - 1)
        elif action_id == constants.LIST_FORMS_NEXT_PAGE:
            current_page = int(action['block_id'].split(':')[-1])  # Extract the current page from the block_id
            result = forms_service.list_forms_command(user, response_url, current_page + 1)
        elif action_id == constants.LIST_FORMS_FIRST_PAGE:
            result = forms_service.list_forms_command(user, response_url, 1)
        elif action_id == constants.LIST_FORMS_LAST_PAGE:
            total_forms = SlackForm.objects(team_id=user.team_id).count()
            last_page = math.ceil(total_forms / constants.FORM_ITEMS_PER_PAGE)
            result = forms_service.list_forms_command(user, response_url, last_page)

    if result:
        return Response(response=json.dumps(result), status=200, mimetype="application/json")
    return Response(status=200, mimetype="application/json")


@app.route('/forms', methods=['GET'])
@user_logged_in
def forms_view():
    page = int(request.args.get('page', 1))
    per_page = min(int(request.args.get('per_page', 10)), 100)
    team_id = session['user_data']['team_id']
    total = SlackForm.objects(team_id=team_id).count()
    page_forms = SlackForm.objects(team_id=team_id).skip((page - 1) * per_page).limit(per_page)
    return render_template('forms.html', forms=page_forms, page=page, per_page=per_page,
                           total=total, SLASH_COMMAND=SLASH_COMMAND, navs=[dict(title='All Forms')])


@app.route('/submissions', methods=['GET'])
@user_logged_in
@form_visible_to_user
def submissions_view():
    # noinspection PyTypeHints
    request.slack_form: SlackForm
    form_id = str(request.slack_form.id)
    form = request.slack_form
    page = int(request.args.get('page', 1))
    per_page = min(int(request.args.get('per_page', 10)), 100)

    total = Submission.objects.filter(form_id=form_id).count()
    page_submissions = Submission.objects.filter(form_id=form_id).skip((page - 1) * per_page).limit(per_page)

    return render_template('submissions.html', form_id=form_id, submissions=page_submissions, page=page,
                           per_page=per_page, total=total, SLASH_COMMAND=SLASH_COMMAND,
                           navs=[dict(title='All Forms', path='/forms'), dict(title=form.name)])


@app.route('/submissions/export/csv')
@user_logged_in
@form_visible_to_user
def export_submissions_csv():
    # noinspection PyUnresolvedReferences
    form_id = request.slack_form.id
    form = SlackForm.objects.get(id=form_id)
    # noinspection PyShadowingNames
    submissions = Submission.objects(form_id=form_id)

    csv_data = io.StringIO()
    writer = csv.writer(csv_data)
    writer.writerow(['submitted by', 'date'] + [field.title for field in form.fields])  # Columns

    for submission in submissions:
        row = [submission.user_name, f"{submission.formatted_date} {submission.formatted_time}"]
        field_values = {field.title: field.value for field in submission.fields}
        row.extend([field_values.get(field.title, '') for field in form.fields])
        writer.writerow(row)

    csv_data.seek(0)
    csv_bytes = io.BytesIO(csv_data.getvalue().encode())

    return send_file(csv_bytes, mimetype='text/csv', as_attachment=True, download_name=f'{form.name}.csv')


@app.route('/submissions/export/xlsx')
@user_logged_in
@form_visible_to_user
def export_submissions_xlsx():
    # noinspection PyUnresolvedReferences
    form_id = request.slack_form.id
    form = SlackForm.objects.get(id=form_id)
    # noinspection PyShadowingNames
    submissions = Submission.objects(form_id=form_id)

    wb = openpyxl.Workbook()
    ws = wb.active

    # Write headers
    headers = ['submitted by', 'date'] + [field.title for field in form.fields]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header

    # Write data
    for row_num, submission in enumerate(submissions, 2):
        row_data = [submission.user_name, f"{submission.formatted_date} {submission.formatted_time}"]
        field_values = {field.title: field.value for field in submission.fields}
        row_data.extend([field_values.get(field.title, '') for field in form.fields])

        for col_num, cell_value in enumerate(row_data, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}{row_num}'] = cell_value

    xlsx_data = io.BytesIO()
    wb.save(xlsx_data)
    xlsx_data.seek(0)
    return send_file(xlsx_data, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'{form.name}.xlsx')


_last_db_healthcheck = None


@app.route('/health')
def health_view():
    global _last_db_healthcheck
    if _last_db_healthcheck is None or _last_db_healthcheck < datetime.now() - timedelta(minutes=1):
        SlackForm.objects().first()  # check that the DB is not down
        _last_db_healthcheck = datetime.now()
    return make_response("OK", 200)
